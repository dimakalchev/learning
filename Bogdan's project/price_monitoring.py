import re
import pandas as pd
import openpyxl
import os
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import datetime

# mydataset = {
#         'cars' : ["BMW", "Volvo", "Ford"],
#         'passings' : [3, 7, 2]
#         }
# myvar = pd.DataFrame(mydataset)
#
# diet = {
#         'calories': ["420", "380", "400"],
#         'duration': ["50", "45", "40"]
# }
# mydiet = pd.DataFrame(diet)
# df = pd.read_csv('file.csv')
# df2 = df.to_dict()

book = openpyxl.open('file.xlsx', read_only=True)

sheet = book.active
first_link = sheet[2][0].value
first_min_price = sheet[2][1].value
second_max_price = sheet[2][2].value

# results = openpyxl.Workbook()
# sheet1 = results.active


# for row in range(1, sheet.max_row + 1):
#     link = sheet[row][0].value
#     min_price = sheet[row][1].value
#     max_price = sheet[row][2].value
#     print(row, link, min_price, max_price)
class Monitoring(webdriver.Chrome):
    def __init__(self, driver_path=r"C:/SeleniumDrivers/chromedriver", teardown=False):
        self.teardown = teardown
        self.driver_path = driver_path
        os.environ['PATH'] += self.driver_path
        super(Monitoring, self).__init__()
        self.implicitly_wait(20)
        self.maximize_window()
        self.results = openpyxl.Workbook()
        self.sheet1 = self.results.active


    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.teardown:
            self.quit()

    def land_first_page(self):
        self.get(first_link)

    def filtres(self):
        min_price_input = self.find_element(By.ID, 'minPrice_')
        min_price_input.send_keys(first_min_price)
        max_price_input = self.find_element(By.ID, 'maxPrice_')
        max_price_input.send_keys(sheet[2][2].value)
        time.sleep(3)
        search = self.find_element(By.CSS_SELECTOR,
                                   'a[href="https://m.ua/ua/m1_magazilla.php?katalog_=206&pf_=1&minPrice_=10000&maxPrice_=20000&sc_id_=980&order_=pop&save_podbor_=1"]')
        search.click()
        filtres = self.find_element(By.ID, 'order_label')
        filtres.click()
        time.sleep(1)
        from_cheapest = self.find_element(By.XPATH,
                                          '/html/body/div[1]/table[3]/tbody/tr/td[3]/table[1]/tbody/tr/td[1]/div[2]/div[1]/a[2]')
        from_cheapest.click()

    def create_book(self):
        book = openpyxl.Workbook()

        sheet = book.active
        sheet['A2'] = 242
        sheet['B3'] = 'hello'

        sheet[1][0].value = 'world'
        sheet.cell(row=1, column=1).value = 'HelloBabies'

        book.save('my_book.xlsx')
        book.close()

    def take_data(self):

        string_with_models_number = self.find_element(By.CLASS_NAME, 'num-models')
        string_models_number = string_with_models_number.get_attribute('textContent')
        print(string_models_number)
        num_models = re.findall('\d+', string_models_number)
        number_of_models = int(num_models[0])
        pages = number_of_models // 15
        print(pages)
        actual_cell = 1
        for x in range(pages+1):
            print(actual_cell)
            block = self.find_element(By.XPATH, '//*[@id="list"]/tbody')
            blocks = block.find_elements(By.CLASS_NAME, 'list-tr')
            new_cell = actual_cell
            for camera_title in blocks:
                all_cameras_titles = camera_title.find_elements(By.CLASS_NAME, 'list-model-title')

                for camera_href in all_cameras_titles:
                    href_blocks = camera_href.find_elements(By.TAG_NAME, 'a')
                    for href in href_blocks:
                        one_href = href.get_attribute('href')
                        self.sheet1.cell(row=new_cell, column=1).value = datetime.date.today()
                        self.sheet1.cell(row=new_cell, column=2).value = new_cell
                        self.sheet1.cell(row=new_cell, column=3).value = one_href
                        self.results.save('results.xlsx')
                        self.results.close()
                        new_cell += 1
                        print(new_cell)
            new_cell = actual_cell

            for camera_lowest_price in blocks:
                all_cameras_lowest_prices = camera_lowest_price.find_elements(By.CLASS_NAME, 'list-big-price')

                for camera_lowest_prices in all_cameras_lowest_prices:
                    lowest_prices_upper = camera_lowest_prices.find_elements(By.CSS_SELECTOR, 'a[title="Порівняти ціни! "]')

                    for camera_lowest_price_spans in lowest_prices_upper:
                        lowest_price_blocks = camera_lowest_price_spans.find_elements(By.TAG_NAME, 'span')
                        for lowest_price in lowest_price_blocks:
                            one_span = lowest_price.get_attribute('textContent')
                            self.sheet1.cell(row=new_cell, column=4).value = one_span
                            self.results.save('results.xlsx')
                            self.results.close()
                            new_cell +=1
            new_cell = actual_cell
            for cameras_description in blocks:
                all_cameras_description = cameras_description.find_elements(By.CLASS_NAME, 'list-model-desc')
                for description in all_cameras_description:

                    one_description = description.get_attribute('textContent')
                    self.sheet1.cell(row=new_cell, column=5).value = one_description
                    self.results.save('results.xlsx')
                    self.results.close()
                    new_cell += 1
            if x >= 0:
                actual_cell += 15
                self.get(f'https://m.ua/ua/m1_magazilla.php?katalog_=206&page_={x+1}&minPrice_=10000&maxPrice_=20000&order_=price')
        print("Information was received successfully")
    # def num_pages(self):
    #     string_with_models_number = self.find_element(By.CLASS_NAME, 'num-models')
    #     string_models_number = string_with_models_number.get_attribute('textContent')
    #     print(string_models_number)
    #     num_models = re.findall('\d+', string_models_number)
    #     number_of_models = int(num_models[0])
    #     pages = number_of_models // 15
    #     print(pages)

    def check_sale(self):
        line=1
        self.get(self.sheet1['C2'].value)
        time.sleep(2)
        price_comparison = self.find_element(By.CLASS_NAME, 'ib small-compare-link-1')
        price_comparison.click()
        sort_price = self.find_element(By.CLASS_NAME, 'sort ')
        sort_price.click()
        lowest_price = self.find_element(By.XPATH, '/html/body/div[1]/table[2]/tbody/tr/td[3]/div[1]/table[5]/tbody/tr[1]/td[3]/a')
        last_price = lowest_price.get_attribute('textContent')
        print(last_price)
        if last_price <= self.sheet1[line][3].value*0.9:
            self.sheet1.cell(row=line,column=0).value= datetime.date.today()
            self.sheet1.cell(row=line, column=3).value = last_price
            print(f"{line}Price is lower")
        line+=1


